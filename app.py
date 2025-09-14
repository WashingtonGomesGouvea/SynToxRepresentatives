from __future__ import annotations

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from typing import Dict, List, Tuple

from config import get_current_datetime, DEFAULT_YEAR, DEFAULT_ACTIVITY_WINDOW_DAYS
from data_loader import load_csvs, enrich_labs_with_reps, merge_gatherings_with_labs
from analytics import (
    filter_active_gatherings,
    compute_credenciamento,
    compute_coleta_status,
    aggregate_volumes,
    compute_kpis,
    build_rankings,
    compute_representative_metrics,
    compute_new_accreditations,
    compute_inactive_labs_alert,
    compute_category_summary,
    compute_geographic_metrics,
    compute_city_metrics,
)
from ui_components import (
    kpi_cards, line_chart_monthly, line_chart_weekly, table,
    performance_dashboard, representative_table, geographic_dashboard
)


# Funções novas para análise de quedas e por representante
# Mudança: Adicionadas para atender à análise de quedas mensais e por rep individual.
# Justificativa: Vetorizado com Pandas para performance; escalável para mil+ labs.
def compute_monthly_variations(monthly: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula variações percentuais entre meses e destaca quedas.
    """
    if monthly.empty:
        return pd.DataFrame()
    
    monthly_sorted = monthly.sort_values('month')
    monthly_sorted['previous_volume'] = monthly_sorted['Volume'].shift(1)  # Vetorizado: shift() é O(1) por grupo.
    monthly_sorted['variation'] = (monthly_sorted['Volume'] - monthly_sorted['previous_volume']) / monthly_sorted['previous_volume'] * 100
    monthly_sorted['variation'] = monthly_sorted['variation'].fillna(0)
    monthly_sorted['is_drop'] = monthly_sorted['variation'] < -10  # Queda significativa >10%
    return monthly_sorted


def compute_rep_accreditations(df_labs_status: pd.DataFrame, rep_name: str, current_date: datetime) -> Dict:
    """
    Calcula credenciamentos e descredenciamentos para um representante específico.
    """
    rep_labs = df_labs_status[df_labs_status['name_rep'] == rep_name]
    
    # Credenciados (ativos)
    credenciados = rep_labs[rep_labs['is_credenciado']]
    num_credenciados = len(credenciados)
    
    # Novos credenciamentos (últimos 3 meses)
    new_accred = compute_new_accreditations(rep_labs, current_date, 3)
    num_new = len(new_accred)
    
    # Descredenciados (não ativos)
    descredenciados = rep_labs[~rep_labs['is_credenciado']]
    num_descred = len(descredenciados)
    
    return {
        'num_credenciados': num_credenciados,
        'num_new_accred': num_new,
        'num_descred': num_descred,
        'new_accred_df': new_accred,
        'descred_df': descredenciados,
        'cred_df': credenciados
    }


def compute_rep_lab_status(df_labs_status: pd.DataFrame, df_gatherings_active: pd.DataFrame, rep_name: str, activity_window: int) -> Dict:
    """
    Calcula status de atividade para labs de um representante.
    """
    rep_labs_ids = df_labs_status[df_labs_status['name_rep'] == rep_name]['_id'].unique()
    rep_gatherings = df_gatherings_active[df_gatherings_active['_laboratory'].isin(rep_labs_ids)]
    rep_labs_status, ativos, inativos, _ = compute_coleta_status(
        df_labs_status[df_labs_status['name_rep'] == rep_name], rep_gatherings, get_current_datetime(), activity_window
    )
    return {
        'ativos': ativos,
        'inativos': inativos,
        'status_df': rep_labs_status
    }


def detect_lab_drops(df_gatherings_active: pd.DataFrame, rep_name: str, target_month: str | None = None) -> pd.DataFrame:
    """
    Detecta quedas bruscas em labs de um representante.
    """
    rep_labs_ids = df_gatherings_active[df_gatherings_active['name_rep'] == rep_name]['_laboratory'].unique()
    rep_gatherings = df_gatherings_active[df_gatherings_active['_laboratory'].isin(rep_labs_ids)]
    
    if rep_gatherings.empty:
        return pd.DataFrame()
    
    # Volumes mensais por lab
    rep_gatherings_copy = rep_gatherings.copy()
    rep_gatherings_copy['month'] = rep_gatherings_copy['createdAt'].dt.to_period('M')
    lab_monthly = rep_gatherings_copy.groupby(['_laboratory', 'month']).size().reset_index(name='Volume')
    lab_monthly = lab_monthly.sort_values(['_laboratory', 'month'])
    
    lab_monthly['previous_volume'] = lab_monthly.groupby('_laboratory')['Volume'].shift(1)
    lab_monthly['variation'] = (lab_monthly['Volume'] - lab_monthly['previous_volume']) / lab_monthly['previous_volume'] * 100
    lab_monthly['variation'] = lab_monthly['variation'].fillna(0)
    lab_monthly['is_drop'] = lab_monthly['variation'] < -30  # Queda brusca >30%
    
    drops = lab_monthly[lab_monthly['is_drop']].copy()
    # Filtrar o mês desejado APÓS calcular as variações (precisa do mês anterior para detectar a queda)
    if target_month is not None and not drops.empty:
        try:
            drops = drops[drops['month'] == pd.Period(target_month)]
        except Exception:
            pass
    if drops.empty:
        return pd.DataFrame()
    
    # Enriquecer com informações úteis para o gestor
    if 'fantasyName' in df_gatherings_active.columns:
        lab_info = df_gatherings_active[['_laboratory', 'fantasyName', 'cnpj']].drop_duplicates()
        drops = drops.merge(lab_info, on='_laboratory', how='left')
        drops['lab_info'] = drops['fantasyName'].fillna('Lab sem nome') + ' (' + drops['cnpj'].fillna('CNPJ não informado') + ')'
    else:
        drops['lab_info'] = 'Laboratório ID: ' + drops['_laboratory'].astype(str)
    
    return drops


def line_chart_with_variations(monthly: pd.DataFrame):
    """
    Gráfico mensal com destaque para variações e quedas.
    """
    st.subheader("📊 Volume Mensal com Variações")
    
    if monthly.empty:
        st.info("Nenhum dado mensal disponível.")
        return
    
    # Condensar quando houver múltiplas categorias (evita linhas duplicadas)
    if 'Categoria' in monthly.columns:
        base = monthly.groupby('month', as_index=False)['Volume'].sum()
    else:
        base = monthly.copy()
    monthly_variations = compute_monthly_variations(base)
    
    fig = go.Figure()
    
    # Linha principal de volume
    fig.add_trace(go.Scatter(
        x=monthly_variations['month'],
        y=monthly_variations['Volume'],
        mode='lines+markers',
        name='Volume de Coletas',
        line=dict(color='#1f77b4', width=3),
        marker=dict(size=8)
    ))
    
    # Adicionar anotações para variações
    for i in range(1, len(monthly_variations)):
        var = monthly_variations.iloc[i]['variation']
        color = 'red' if var < 0 else 'green'
        fig.add_annotation(
            x=monthly_variations.iloc[i]['month'],
            y=monthly_variations.iloc[i]['Volume'],
            text=f"{var:.1f}%",
            showarrow=True,
            arrowhead=1,
            ax=0,
            ay=-40 if var < 0 else 40,
            font=dict(color=color)
        )
    
    fig.update_layout(
        title="Volume Mensal com Variações Percentuais",
        xaxis_title="Mês",
        yaxis_title="Quantidade de Coletas",
        hovermode='x unified'
    )
    
    fig.update_traces(
        hovertemplate="<b>Volume</b>: %{y:,.0f}<br>Variação: %{text:.1f}%",
        text=monthly_variations['variation']
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Tabela de variações
    st.subheader("📋 Tabela de Variações Mensais")
    display_df = monthly_variations[['month', 'Volume', 'variation']].copy()
    display_df['variation'] = display_df['variation'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
    display_df = display_df.rename(columns={'month': 'Mês', 'Volume': 'Coletas', 'variation': 'Variação (%)'})
    st.dataframe(display_df, use_container_width=True)


def create_rep_charts(rep_name: str, df_gatherings_active: pd.DataFrame, df_labs_status: pd.DataFrame, activity_window: int):
    """
    Cria gráficos para análise do representante.
    """
    charts = {}
    
    # Filtrar dados do representante
    rep_labs_ids = df_labs_status[df_labs_status['name_rep'] == rep_name]['_id'].unique()
    rep_gatherings = df_gatherings_active[df_gatherings_active['_laboratory'].isin(rep_labs_ids)]
    
    if not rep_gatherings.empty:
        # 1. Gráfico de volume mensal
        rep_gatherings['month'] = rep_gatherings['createdAt'].dt.to_period('M')
        monthly_volume = rep_gatherings.groupby('month').size().reset_index(name='Volume')
        monthly_volume['month'] = monthly_volume['month'].astype(str)
        
        fig_monthly = go.Figure()
        fig_monthly.add_trace(go.Scatter(
            x=monthly_volume['month'],
            y=monthly_volume['Volume'],
            mode='lines+markers+text',
            name='Volume de Coletas',
            line=dict(color='#1f77b4', width=3),
            marker=dict(size=8),
            text=monthly_volume['Volume'],
            textposition='top center',
            textfont=dict(size=10, color='#1f77b4')
        ))
        fig_monthly.update_layout(
            title=f"Volume Mensal de Coletas - {rep_name}",
            xaxis_title="Mês",
            yaxis_title="Quantidade de Coletas",
            height=400
        )
        charts['monthly_volume'] = fig_monthly
        
        # 2. Gráfico de performance por laboratório
        lab_performance = rep_gatherings.groupby('fantasyName').size().reset_index(name='Volume')
        lab_performance = lab_performance.sort_values('Volume', ascending=True).tail(10)  # Top 10
        
        fig_labs = go.Figure()
        fig_labs.add_trace(go.Bar(
            x=lab_performance['Volume'],
            y=lab_performance['fantasyName'],
            orientation='h',
            marker_color='#2ca02c',
            text=lab_performance['Volume'],
            textposition='auto',
            textfont=dict(size=10, color='white')
        ))
        fig_labs.update_layout(
            title=f"Top 10 Laboratórios por Volume - {rep_name}",
            xaxis_title="Quantidade de Coletas",
            yaxis_title="Laboratório",
            height=500
        )
        charts['lab_performance'] = fig_labs
        
        # 3. Gráfico de status dos laboratórios (ATIVIDADE DE COLETA)
        rep_labs = df_labs_status[df_labs_status['name_rep'] == rep_name]
        # Calcular status de atividade usando a função existente
        rep_labs_with_status, ativos, inativos, _ = compute_coleta_status(
            rep_labs, rep_gatherings, get_current_datetime(), activity_window
        )
        status_counts = rep_labs_with_status['ativo_coleta'].value_counts()
        
        fig_status = go.Figure()
        fig_status.add_trace(go.Pie(
            labels=['Ativos em Coleta', 'Inativos em Coleta'],
            values=[status_counts.get(True, 0), status_counts.get(False, 0)],
            hole=0.4,
            marker_colors=['#2ca02c', '#d62728'],
            textinfo='none'
        ))
        fig_status.update_layout(
            title=f"Status de Atividade de Coleta - {rep_name}<br><sub>Labs que coletaram nos últimos {activity_window} dias</sub>",
            height=400
        )
        charts['status_pie'] = fig_status
        
        # 4. Gráfico de tendência semanal
        rep_gatherings['week'] = rep_gatherings['createdAt'].dt.to_period('W')
        weekly_volume = rep_gatherings.groupby('week').size().reset_index(name='Volume')
        weekly_volume['week'] = weekly_volume['week'].astype(str)
        
        fig_weekly = go.Figure()
        fig_weekly.add_trace(go.Scatter(
            x=weekly_volume['week'],
            y=weekly_volume['Volume'],
            mode='lines+markers+text',
            name='Volume Semanal',
            line=dict(color='#ff7f0e', width=2),
            marker=dict(size=6),
            text=weekly_volume['Volume'],
            textposition='top center',
            textfont=dict(size=9, color='#ff7f0e')
        ))
        fig_weekly.update_layout(
            title=f"Tendência Semanal - {rep_name}",
            xaxis_title="Semana",
            yaxis_title="Quantidade de Coletas",
            height=400
        )
        charts['weekly_trend'] = fig_weekly
    
    return charts


def generate_rep_pdf(rep_name: str, accred_metrics: dict, status_metrics: dict, charts: dict, current_date: datetime):
    """
    Gera relatório PDF avançado para o representante com gráficos e análises detalhadas.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')
    import io
    import os
    
    # Configurar estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Centralizado
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        textColor=colors.darkblue
    )
    
    # Criar documento
    filename = f"relatorio_{rep_name.replace(' ', '_')}_{current_date.strftime('%Y%m%d_%H%M')}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=A4)
    story = []
    
    # Título principal
    story.append(Paragraph(f"Relatório de Performance - {rep_name}", title_style))
    story.append(Paragraph(f"Data do Relatório: {current_date.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Resumo Executivo
    story.append(Paragraph("📊 Resumo Executivo", subtitle_style))
    
    # Métricas em tabela
    total_labs = accred_metrics['num_credenciados']
    activation_rate = (status_metrics['ativos'] / total_labs) * 100 if total_labs > 0 else 0
    
    metrics_data = [
        ['Métrica', 'Valor', 'Status'],
        ['Labs Credenciados', str(accred_metrics['num_credenciados']), '✅'],
        ['Labs Ativos', str(status_metrics['ativos']), '🟢'],
        ['Labs Inativos', str(status_metrics['inativos']), '🔴'],
        ['Taxa de Ativação', f"{activation_rate:.1f}%", '📊'],
        ['Novos Credenciamentos (3m)', str(accred_metrics['num_new_accred']), '🆕'],
        ['Labs Descredenciados', str(accred_metrics['num_descred']), '❌']
    ]
    
    metrics_table = Table(metrics_data, colWidths=[2*inch, 1.5*inch, 0.5*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 20))
    
    # Gráfico de Status (Pizza)
    if 'status_pie' in charts:
        story.append(Paragraph("🥧 Status de Atividade", subtitle_style))
        
        # Criar gráfico de pizza com matplotlib
        fig, ax = plt.subplots(figsize=(8, 6))
        status_counts = [status_metrics['ativos'], status_metrics['inativos']]
        labels = ['Ativos', 'Inativos']
        colors_pie = ['#2ca02c', '#d62728']
        
        ax.pie(status_counts, labels=labels, colors=colors_pie, autopct='%1.1f%%', startangle=90)
        ax.set_title('Distribuição de Status dos Laboratórios')
        
        # Salvar gráfico temporariamente
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        # Adicionar ao PDF
        img = Image(img_buffer)
        img.drawHeight = 3*inch
        img.drawWidth = 4*inch
        story.append(img)
        story.append(Spacer(1, 20))
    
    # Análise de Performance
    story.append(Paragraph("📈 Análise de Performance", subtitle_style))
    
    # Calcular insights
    insights = []
    if status_metrics['inativos'] > 0:
        inactive_percentage = (status_metrics['inativos'] / total_labs) * 100 if total_labs > 0 else 0
        insights.append(f"• {status_metrics['inativos']} labs inativos ({inactive_percentage:.1f}% do total)")
    
    if accred_metrics['num_descred'] > 0:
        insights.append(f"• {accred_metrics['num_descred']} labs descredenciados")
    
    if accred_metrics['num_new_accred'] == 0:
        insights.append("• Nenhum novo credenciamento nos últimos 3 meses")
    elif accred_metrics['num_new_accred'] > 0:
        insights.append(f"• {accred_metrics['num_new_accred']} novos credenciamentos (crescimento positivo)")
    
    if activation_rate < 70:
        insights.append(f"• Taxa de ativação baixa ({activation_rate:.1f}%) - atenção necessária")
    elif activation_rate >= 80:
        insights.append(f"• Taxa de ativação excelente ({activation_rate:.1f}%)")
    
    for insight in insights:
        story.append(Paragraph(insight, styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Prioridades de Ação
    story.append(Paragraph("🎯 Prioridades de Ação", subtitle_style))
    
    priorities = []
    if status_metrics['inativos'] > 0:
        priorities.append("1. Contatar labs inativos para entender motivos da inatividade")
        priorities.append("2. Planejar visitas aos labs críticos (mais de 30 dias sem coleta)")
        priorities.append("3. Implementar programa de reativação")
    
    if accred_metrics['num_new_accred'] == 0:
        priorities.append("4. Focar em novos credenciamentos")
        priorities.append("5. Identificar oportunidades de mercado")
    
    if activation_rate < 70:
        priorities.append("6. Revisar estratégia de relacionamento")
        priorities.append("7. Implementar programa de ativação")
    
    if accred_metrics['num_descred'] > 0:
        priorities.append("8. Investigar motivos dos descredenciamentos")
    
    for priority in priorities:
        story.append(Paragraph(priority, styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Top Labs para Foco
    if 'lab_performance' in charts:
        story.append(Paragraph("🏆 Top Laboratórios para Foco", subtitle_style))
        
        # Extrair dados do gráfico de performance
        try:
            # Aqui você pode adicionar uma tabela com os top labs
            story.append(Paragraph("• Focar nos laboratórios com maior volume de coletas", styles['Normal']))
            story.append(Paragraph("• Identificar oportunidades de crescimento nos labs de médio volume", styles['Normal']))
            story.append(Paragraph("• Reativar labs inativos com histórico de alto volume", styles['Normal']))
        except:
            story.append(Paragraph("Dados de performance dos laboratórios disponíveis no dashboard online.", styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Recomendações Estratégicas
    story.append(Paragraph("💡 Recomendações Estratégicas", subtitle_style))
    
    recommendations = []
    if status_metrics['inativos'] > status_metrics['ativos']:
        recommendations.append("• Implementar programa de reativação urgente")
        recommendations.append("• Criar equipe dedicada para contato com labs inativos")
        recommendations.append("• Oferecer incentivos para retomada de atividades")
    
    if accred_metrics['num_new_accred'] < 5:
        recommendations.append("• Intensificar esforços de prospecção")
        recommendations.append("• Participar de eventos do setor")
        recommendations.append("• Desenvolver material promocional específico")
    
    if activation_rate < 60:
        recommendations.append("• Revisar processo de onboarding de novos labs")
        recommendations.append("• Implementar acompanhamento mais frequente")
        recommendations.append("• Criar programa de suporte técnico")
    
    for rec in recommendations:
        story.append(Paragraph(rec, styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Footer
    story.append(Paragraph("---", styles['Normal']))
    story.append(Paragraph("Relatório gerado automaticamente pelo Sistema de Gestão de Representantes", styles['Normal']))
    story.append(Paragraph("Para mais detalhes, acesse o dashboard online.", styles['Normal']))
    
    # Gerar PDF
    doc.build(story)
    
    # Limpar arquivos temporários
    if 'img_buffer' in locals():
        img_buffer.close()
    
    return filename


def generate_executive_pdf(rep_name: str, accred_metrics: dict, status_metrics: dict, charts: dict, current_date: datetime, top_labs_df: pd.DataFrame | None = None, inactives_df: pd.DataFrame | None = None, drops_df: pd.DataFrame | None = None):
    """
    Gera PDF executivo focado em ações e resumo.
    """
    from fpdf import FPDF
    import os
    try:
        from matplotlib import font_manager as _fm
    except Exception:
        _fm = None
    
    def _sanitize(text: str) -> str:
        if text is None:
            return ""
        # Mantém acentos; apenas normaliza aspas/dashes problemáticos
        replacements = {
            "…": "...",
        }
        t = str(text)
        for k, v in replacements.items():
            t = t.replace(k, v)
        return t

    def cell(w, h, txt, **kwargs):
        return pdf.cell(w, h, _sanitize(str(txt)), **kwargs)

    import re
    pdf = FPDF()
    # Margens e quebra automática
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    # Garantir fonte Unicode embutida (DejaVu/Arial)
    def _ensure_unicode_font(p: FPDF) -> str:
        font_name = "Unicode"
        # Já registrada?
        try:
            p.set_font(font_name, size=11)
            return font_name
        except Exception:
            pass
        candidates: list[tuple[str, str | None]] = []
        base_dir = os.path.dirname(__file__) if '__file__' in globals() else os.getcwd()
        local_reg = os.path.join(base_dir, 'fonts', 'DejaVuSans.ttf')
        local_bold = os.path.join(base_dir, 'fonts', 'DejaVuSans-Bold.ttf')
        candidates.append((local_reg, local_bold))
        # Matplotlib (geralmente presente)
        try:
            if _fm is not None:
                reg = _fm.findfont('DejaVu Sans', fallback_to_default=True)
                bold = _fm.findfont('DejaVu Sans:style=bold', fallback_to_default=True)
                candidates.append((reg, bold))
        except Exception:
            pass
        # Windows Arial
        if os.name == 'nt':
            windir = os.environ.get('WINDIR', 'C:\\Windows')
            arial = os.path.join(windir, 'Fonts', 'arial.ttf')
            arialb = os.path.join(windir, 'Fonts', 'arialbd.ttf')
            candidates.append((arial, arialb))
        # Linux comum
        candidates.append(('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))

        for reg, bold in candidates:
            try:
                if reg and os.path.exists(reg):
                    p.add_font(font_name, '', reg, uni=True)
                    if bold and os.path.exists(bold):
                        p.add_font(font_name, 'B', bold, uni=True)
                    else:
                        # Se não há bold, usa regular mesmo
                        p.add_font(font_name, 'B', reg, uni=True)
                    p.set_font(font_name, size=11)
                    return font_name
            except Exception:
                continue
        # Fallback: usa Arial core; manterá sanitização leve
        return 'Arial'

    _unicode_font_name = _ensure_unicode_font(pdf)

    # Dimensões úteis
    page_width = pdf.w - 2 * pdf.l_margin

    # Helpers de formatação
    def _digits_only(text: str) -> str:
        return re.sub(r"\D+", "", str(text or ""))

    def _format_cnpj(cnpj_val: str) -> str:
        digits = _digits_only(cnpj_val)
        if len(digits) == 14:
            return f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"
        return digits or "-"

    def _ellipsize(text: str, max_w: float) -> str:
        text = _sanitize(text)
        if max_w <= 10:
            return ""
        if pdf.get_string_width(text) <= max_w:
            return text
        ellipsis = "..."
        max_w = max(10, max_w - pdf.get_string_width(ellipsis))
        result = text
        while result and pdf.get_string_width(result) > max_w:
            result = result[:-1]
        return result + ellipsis

    def _lab_label(name: str, cnpj_val: str | None, max_w: float) -> str:
        cnpj_fmt = _format_cnpj(cnpj_val) if cnpj_val else None
        base_name = str(name or "").strip()
        if "(" in base_name and ")" in base_name:
            try:
                base_name = base_name.split("(")[0].strip()
            except Exception:
                pass
        prefix = f"CNPJ {cnpj_fmt} - " if cnpj_fmt else ""
        return _ellipsize(prefix + base_name, max_w)

    # Helpers de layout
    def draw_section_header(title: str):
        # Garante início no alinhamento da margem esquerda
        pdf.set_x(pdf.l_margin)
        pdf.set_fill_color(34, 74, 120)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 13)
        cell(page_width, 9, title, ln=True, align='L', border=0, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

    def multi_cell(w, h, txt, **kwargs):
        return pdf.multi_cell(w, h, _sanitize(str(txt)), **kwargs)

    # Configurar fonte e cabeçalho
    pdf.set_font(_unicode_font_name, "B", 16)
    cell(0, 10, f"Resumo Executivo - {rep_name}", ln=True, align='C')
    pdf.set_draw_color(180, 180, 180)
    pdf.line(pdf.l_margin, pdf.get_y() + 1, pdf.w - pdf.r_margin, pdf.get_y() + 1)
    pdf.ln(4)

    # Data do relatório
    pdf.set_font(_unicode_font_name, "", 11)
    cell(0, 8, f"Data: {current_date.strftime('%d/%m/%Y %H:%M')}", ln=True)
    pdf.ln(4)
    
    # Métricas principais - tabela limpa
    draw_section_header("Métricas Principais")
    total_labs = accred_metrics['num_credenciados']
    activation_rate = (status_metrics['ativos'] / total_labs) * 100 if total_labs > 0 else 0

    metrics = [
        ("Labs Credenciados", f"{accred_metrics['num_credenciados']}"),
        ("Labs Ativos", f"{status_metrics['ativos']} ({activation_rate:.1f}%)"),
        ("Labs Inativos", f"{status_metrics['inativos']}"),
        ("Novos Credenciamentos (3m)", f"{accred_metrics['num_new_accred']}"),
        ("Labs Descredenciados", f"{accred_metrics['num_descred']}")
    ]

    # Larguras robustas: garante espaço mínimo para o valor
    col_value = max(40, page_width * 0.33)
    col_label = page_width - col_value

    # Cabeçalho
    pdf.set_fill_color(240, 240, 240)
    pdf.set_draw_color(150, 150, 150)
    pdf.set_line_width(0.2)
    pdf.set_font(_unicode_font_name, "B", 11)
    cell(col_label, 8, "Métrica", border=1, fill=True)
    cell(col_value, 8, "Valor", border=1, ln=True, fill=True, align='R')
    pdf.set_font(_unicode_font_name, "", 11)
    # Linhas com zebra
    for idx, (label, value) in enumerate(metrics):
        if idx % 2 == 0:
            pdf.set_fill_color(250, 250, 250)
        else:
            pdf.set_fill_color(245, 245, 245)
        cell(col_label, 8, label, border=1, fill=True)
        cell(col_value, 8, value, border=1, ln=True, fill=True, align='R')
    pdf.ln(6)
    
    # Ações Prioritárias
    draw_section_header("Ações Prioritárias")
    pdf.set_font(_unicode_font_name, "", 11)
    
    actions = []
    if status_metrics['inativos'] > 0:
        actions.append(f"1. Contatar {status_metrics['inativos']} labs inativos")
        actions.append("2. Planejar visitas aos labs criticos")
        actions.append("3. Implementar programa de reativacao")
    
    if accred_metrics['num_new_accred'] == 0:
        actions.append("4. Focar em novos credenciamentos")
        actions.append("5. Identificar oportunidades de mercado")
    
    if activation_rate < 70:
        actions.append("6. Revisar estrategia de relacionamento")
        actions.append("7. Implementar programa de ativacao")
    
    if accred_metrics['num_descred'] > 0:
        actions.append("8. Investigar motivos dos descredenciamentos")
    
    for action in actions:
        # indentação suave e quebra automática
        pdf.set_x(pdf.l_margin + 4)
        multi_cell(page_width - 8, 7, action)
    pdf.ln(4)
    
    # Observações (linguagem neutra)
    draw_section_header("Observações")
    pdf.set_font(_unicode_font_name, "", 11)
    observations = []
    if activation_rate >= 80:
        observations.append("Indicadores positivos. Manter acompanhamento de rotina.")
    elif activation_rate >= 70:
        observations.append("Bom desempenho. Sugerido monitorar labs com menor atividade.")
    elif activation_rate >= 50:
        observations.append("Oportunidade de melhoria. Priorizar reativacao e relacionamento.")
    else:
        observations.append("Recomendado plano dedicado de reativacao e novos credenciamentos.")

    for obs in observations:
        multi_cell(page_width, 7, obs)
    
    # Seções com dados reais
    pdf.ln(8)
    pdf.set_font("Arial", "B", 13)
    cell(0, 8, "Top 10 Laboratórios por Volume", ln=True)
    pdf.set_font("Arial", "", 11)
    if top_labs_df is not None and not top_labs_df.empty:
        # Tamanhos das colunas (com mínimos seguros)
        col_vol = max(35, page_width * 0.20)
        col_lab = max(60, page_width - col_vol)
        # Cabeçalho
        pdf.set_fill_color(230, 230, 230)
        pdf.set_draw_color(80, 80, 80)
        pdf.set_line_width(0.2)
        pdf.set_font(_unicode_font_name, "B", 11)
        cell(col_lab, 8, "Laboratório", border=1, fill=True)
        cell(col_vol, 8, "Volume", border=1, ln=True, fill=True, align='R')
        pdf.set_font(_unicode_font_name, "", 11)
        for _, row in top_labs_df.head(10).iterrows():
            lab_name_raw = str(row.get('Laboratorio', row.get('fantasyName', '')))
            lab_name = _lab_label(lab_name_raw, row.get('cnpj'), col_lab)
            volume = f"{int(row.get('Volume', row.get('Volume de Coletas', 0))):,}".replace(",", ".")
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.set_font(_unicode_font_name, "B", 11)
                cell(col_lab, 8, "Laboratório", border=1, fill=True)
                cell(col_vol, 8, "Volume", border=1, ln=True, fill=True, align='R')
                pdf.set_font(_unicode_font_name, "", 11)
            # Linha
            cell(col_lab, 7, lab_name, border=1)
            cell(col_vol, 7, volume, border=1, ln=True, align='R')
    else:
        cell(0, 8, "Sem dados suficientes para este periodo.", ln=True)

    # Inativos criticos
    pdf.ln(8)
    pdf.set_font(_unicode_font_name, "B", 13)
    cell(0, 8, "Top Labs Inativos (mais dias sem coleta)", ln=True)
    pdf.set_font(_unicode_font_name, "", 11)
    if inactives_df is not None and not inactives_df.empty:
        pdf.set_font(_unicode_font_name, "B", 11)
        pdf.set_fill_color(230, 230, 230)
        # Larguras (com mínimos seguros)
        col_days = max(30, page_width * 0.18)
        col_last = max(50, page_width * 0.32)
        col_lab = max(60, page_width - col_days - col_last)
        cell(col_lab, 8, "Laboratório", border=1, fill=True)
        cell(col_days, 8, "Dias sem coleta", border=1, fill=True, align='R')
        cell(col_last, 8, "Última coleta", border=1, ln=True, fill=True)
        pdf.set_font(_unicode_font_name, "", 11)
        for _, row in inactives_df.head(10).iterrows():
            lab_name = _lab_label(row.get('Laboratorio', row.get('fantasyName', '')), row.get('cnpj'), col_lab)
            dias = str(row.get('Dias sem Coletar', row.get('days_since_last', row.get('days_since_last_display', ''))))
            ultima = str(row.get('Ultima Coleta', row.get('ultima_coleta_str', '')))
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.set_font(_unicode_font_name, "B", 11)
                pdf.set_fill_color(230, 230, 230)
                cell(col_lab, 8, "Laboratório", border=1, fill=True)
                cell(col_days, 8, "Dias sem coleta", border=1, fill=True, align='R')
                cell(col_last, 8, "Última coleta", border=1, ln=True, fill=True)
                pdf.set_font(_unicode_font_name, "", 11)
            cell(col_lab, 7, lab_name, border=1)
            cell(col_days, 7, dias, border=1, align='R')
            cell(col_last, 7, ultima, border=1, ln=True)
    else:
        cell(0, 8, "Sem labs inativos para exibir.", ln=True)

    # Maiores quedas
    pdf.ln(8)
    pdf.set_font(_unicode_font_name, "B", 13)
    cell(0, 8, "Maiores Quedas (mês a mês)", ln=True)
    pdf.set_font(_unicode_font_name, "", 11)
    if drops_df is not None and not drops_df.empty:
        pdf.set_font(_unicode_font_name, "B", 11)
        pdf.set_fill_color(230, 230, 230)
        # Larguras (com mínimos seguros)
        col_mes = max(25, page_width * 0.16)
        col_atual = max(25, page_width * 0.12)
        col_ant = max(25, page_width * 0.14)
        col_queda = max(30, page_width * 0.12)
        col_lab = max(60, page_width - col_mes - col_atual - col_ant - col_queda)
        cell(col_lab, 8, "Laboratório", border=1, fill=True)
        cell(col_mes, 8, "Mês", border=1, fill=True)
        cell(col_atual, 8, "Atual", border=1, fill=True, align='R')
        cell(col_ant, 8, "Anterior", border=1, fill=True, align='R')
        cell(col_queda, 8, "Queda %", border=1, ln=True, fill=True, align='R')
        pdf.set_font(_unicode_font_name, "", 11)
        for _, row in drops_df.head(10).iterrows():
            raw_name = str(row.get('Laboratorio', row.get('lab_info', '')))
            cnpj_candidate = row.get('cnpj') or (re.search(r"\(?([\d\.\/-]{11,})\)?", raw_name).group(1) if re.search(r"\d{11,}", raw_name) else None)
            lab = _lab_label(raw_name, cnpj_candidate, col_lab)
            mes = str(row.get('Mes', row.get('month', '')))
            atual_val = row.get('Coletas Atual', row.get('Volume', 0))
            ant_val = row.get('Coletas Anterior', row.get('previous_volume', 0))
            queda_val = row.get('Queda (%)', row.get('variation', 0.0))
            atual = f"{int(atual_val):,}".replace(",", ".")
            ant = f"{int(ant_val):,}".replace(",", ".")
            try:
                queda_num = float(str(queda_val).replace('%',''))
            except:
                queda_num = float(queda_val) if isinstance(queda_val, (int,float)) else 0.0
            queda = f"{queda_num:.1f}%"
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.set_font(_unicode_font_name, "B", 11)
                pdf.set_fill_color(230, 230, 230)
                cell(col_lab, 8, "Laboratório", border=1, fill=True)
                cell(col_mes, 8, "Mês", border=1, fill=True)
                cell(col_atual, 8, "Atual", border=1, fill=True, align='R')
                cell(col_ant, 8, "Anterior", border=1, fill=True, align='R')
                cell(col_queda, 8, "Queda %", border=1, ln=True, fill=True, align='R')
                pdf.set_font(_unicode_font_name, "", 11)
            cell(col_lab, 7, lab, border=1)
            cell(col_mes, 7, mes, border=1)
            cell(col_atual, 7, atual, border=1, align='R')
            cell(col_ant, 7, ant, border=1, align='R')
            cell(col_queda, 7, str(queda), border=1, ln=True, align='R')
    else:
        cell(0, 8, "Sem quedas significativas detectadas.", ln=True)

    # Salvar PDF
    filename = f"resumo_executivo_{rep_name.replace(' ', '_')}_{current_date.strftime('%Y%m%d_%H%M')}.pdf"
    pdf.output(filename)
    return filename


def generate_complete_excel(rep_name: str, accred_metrics: dict, status_metrics: dict, df_gatherings_active: pd.DataFrame, df_labs_status: pd.DataFrame, activity_window: int, current_date: datetime):
    """
    Gera Excel completo com todos os dados organizados em abas.
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    import os
    
    # Criar workbook
    wb = Workbook()
    
    # Remover aba padrão
    wb.remove(wb.active)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Utilitário: autoajustar colunas, congelar cabeçalho e habilitar filtros
    def _autofit_freeze_and_filter(ws):
        try:
            if ws.max_column == 0 or ws.max_row == 0:
                return
            # Calcular larguras por coluna
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is None:
                        length = 0
                    else:
                        try:
                            if hasattr(value, 'strftime'):
                                text = value.strftime('%d/%m/%Y')
                            else:
                                text = str(value)
                        except Exception:
                            text = str(value)
                        length = len(text)
                    if length > max_length:
                        max_length = length
                adjusted = min(max_length + 2, 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted
            # Congelar cabeçalho e centralizar
            has_header = any(ws.cell(row=1, column=c).value is not None for c in range(1, ws.max_column + 1))
            if has_header:
                ws.freeze_panes = "A2"
                for c in range(1, ws.max_column + 1):
                    hdr = ws.cell(row=1, column=c)
                    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Habilitar filtros para o intervalo utilizado
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        except Exception:
            # Em caso de qualquer erro, não interromper geração do arquivo
            pass
    
    # 1. ABA: Resumo Executivo
    ws_summary = wb.create_sheet("Resumo Executivo")
    
    # Dados do resumo
    total_labs = accred_metrics['num_credenciados']
    activation_rate = (status_metrics['ativos'] / total_labs) * 100 if total_labs > 0 else 0
    
    summary_data = [
        ["Métrica", "Valor", "Status"],
        ["Labs Credenciados", accred_metrics['num_credenciados'], "Total"],
        ["Labs Ativos", status_metrics['ativos'], f"{activation_rate:.1f}%"],
        ["Labs Inativos", status_metrics['inativos'], f"{(100-activation_rate):.1f}%"],
        ["Taxa de Ativação", f"{activation_rate:.1f}%", "Performance"],
        ["Novos Credenciamentos (3m)", accred_metrics['num_new_accred'], "Crescimento"],
        ["Labs Descredenciados", accred_metrics['num_descred'], "Perdas"]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:  # Header
                cell.font = header_font
                cell.fill = header_fill
    
    # 2. ABA: Novos Credenciamentos
    ws_new = wb.create_sheet("Novos Credenciamentos")
    if not accred_metrics['new_accred_df'].empty:
        new_labs_clean = accred_metrics['new_accred_df'].copy()
        useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'data_credenciamento', 'dias_credenciado']
        available_columns = [col for col in useful_columns if col in new_labs_clean.columns]
        
        if available_columns:
            new_labs_clean = new_labs_clean[available_columns]
            column_mapping = {
                'fantasyName': 'Nome do Laboratório',
                'cnpj': 'CNPJ',
                'name_rep': 'Representante',
                'Categoria': 'Tipo',
                'data_credenciamento': 'Data de Credenciamento',
                'dias_credenciado': 'Dias Credenciado'
            }
            new_labs_clean = new_labs_clean.rename(columns=column_mapping)
            
            # Formatar CNPJ
            if 'CNPJ' in new_labs_clean.columns:
                new_labs_clean['CNPJ'] = new_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
            
            # Adicionar dados
            for r in dataframe_to_rows(new_labs_clean, index=False, header=True):
                ws_new.append(r)
            
            # Aplicar estilos
            for row in ws_new.iter_rows(min_row=1, max_row=len(new_labs_clean)+1):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    
    # 3. ABA: Labs Descredenciados
    ws_descred = wb.create_sheet("Labs Descredenciados")
    if not accred_metrics['descred_df'].empty:
        descred_labs_clean = accred_metrics['descred_df'].copy()
        useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'exclusionDate']
        available_columns = [col for col in useful_columns if col in descred_labs_clean.columns]
        
        if available_columns:
            descred_labs_clean = descred_labs_clean[available_columns]
            column_mapping = {
                'fantasyName': 'Nome do Laboratório',
                'cnpj': 'CNPJ',
                'name_rep': 'Representante',
                'Categoria': 'Tipo',
                'exclusionDate': 'Data de Descredenciamento'
            }
            descred_labs_clean = descred_labs_clean.rename(columns=column_mapping)
            
            # Formatar CNPJ
            if 'CNPJ' in descred_labs_clean.columns:
                descred_labs_clean['CNPJ'] = descred_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
            
            # Adicionar dados
            for r in dataframe_to_rows(descred_labs_clean, index=False, header=True):
                ws_descred.append(r)
            
            # Aplicar estilos
            for row in ws_descred.iter_rows(min_row=1, max_row=len(descred_labs_clean)+1):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    
    # 4. ABA: Labs Credenciados
    ws_cred = wb.create_sheet("Labs Credenciados")
    if not accred_metrics['cred_df'].empty:
        cred_labs_clean = accred_metrics['cred_df'].copy()
        useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'data_credenciamento', 'dias_credenciado']
        available_columns = [col for col in useful_columns if col in cred_labs_clean.columns]
        
        if available_columns:
            cred_labs_clean = cred_labs_clean[available_columns]
            column_mapping = {
                'fantasyName': 'Nome do Laboratório',
                'cnpj': 'CNPJ',
                'name_rep': 'Representante',
                'Categoria': 'Tipo',
                'data_credenciamento': 'Data de Credenciamento',
                'dias_credenciado': 'Dias Credenciado'
            }
            cred_labs_clean = cred_labs_clean.rename(columns=column_mapping)
            
            # Formatar CNPJ
            if 'CNPJ' in cred_labs_clean.columns:
                cred_labs_clean['CNPJ'] = cred_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
            
            # Adicionar dados
            for r in dataframe_to_rows(cred_labs_clean, index=False, header=True):
                ws_cred.append(r)
            
            # Aplicar estilos
            for row in ws_cred.iter_rows(min_row=1, max_row=len(cred_labs_clean)+1):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    
    # 5. ABA: Status de Coletas
    ws_status = wb.create_sheet("Status de Coletas")
    if 'status_df' in status_metrics and not status_metrics['status_df'].empty:
        status_labs_clean = status_metrics['status_df'].copy()
        useful_columns = ['fantasyName', 'cnpj', 'ativo_coleta', 'days_since_last_display', 'ultima_coleta_str']
        available_columns = [col for col in useful_columns if col in status_labs_clean.columns]
        
        if available_columns:
            status_labs_clean = status_labs_clean[available_columns]
            column_mapping = {
                'fantasyName': 'Nome do Laboratório',
                'cnpj': 'CNPJ',
                'ativo_coleta': 'Ativo em Coleta',
                'days_since_last_display': 'Dias sem Coletar',
                'ultima_coleta_str': 'Última Coleta'
            }
            status_labs_clean = status_labs_clean.rename(columns=column_mapping)
            
            # Formatar CNPJ
            if 'CNPJ' in status_labs_clean.columns:
                status_labs_clean['CNPJ'] = status_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
            
            # Formatar status
            if 'Ativo em Coleta' in status_labs_clean.columns:
                status_labs_clean['Ativo em Coleta'] = status_labs_clean['Ativo em Coleta'].map({True: 'Sim', False: 'Não'})
            
            # Adicionar dados
            for r in dataframe_to_rows(status_labs_clean, index=False, header=True):
                ws_status.append(r)
            
            # Aplicar estilos
            for row in ws_status.iter_rows(min_row=1, max_row=len(status_labs_clean)+1):
                for cell in row:
                    cell.border = border
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
    
    # 6. ABA: Quedas Bruscas
    ws_drops = wb.create_sheet("Quedas Bruscas")
    drops_df = detect_lab_drops(df_gatherings_active, rep_name)
    if not drops_df.empty and 'variation' in drops_df.columns:
        drops_clean = drops_df.copy()
        # Garantir strings para Excel
        drops_clean['month'] = drops_clean['month'].astype(str)
        drops_clean['variation'] = drops_clean['variation'].apply(lambda x: f"{x:.1f}%")
        drops_clean['Volume'] = drops_clean['Volume'].apply(lambda x: f"{int(x):,}".replace(",", "."))
        
        # Calcular mês anterior para referência
        drops_clean['previous_month'] = drops_clean['month'].apply(lambda x: str(pd.Period(x) - 1))
        drops_clean['previous_volume'] = drops_clean['previous_volume'].apply(lambda x: f"{int(x):,}".replace(",", "."))
        
        # Criar coluna com mês de referência
        drops_clean['previous_volume_with_month'] = drops_clean['previous_volume'] + ' (' + drops_clean['previous_month'] + ')'
        
        display_columns = ['lab_info', 'month', 'Volume', 'previous_volume_with_month', 'variation']
        column_mapping = {
            'lab_info': 'Laboratório',
            'month': 'Mês',
            'Volume': 'Coletas Atual',
            'previous_volume_with_month': 'Coletas Anterior (Mês)',
            'variation': 'Queda (%)'
        }
        
        drops_clean = drops_clean[display_columns]
        drops_clean = drops_clean.rename(columns=column_mapping)
        
        # Adicionar dados
        for r in dataframe_to_rows(drops_clean, index=False, header=True):
            ws_drops.append(r)
        
        # Aplicar estilos
        for row in ws_drops.iter_rows(min_row=1, max_row=len(drops_clean)+1):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
    
    # 7. ABA: Performance por Lab
    ws_performance = wb.create_sheet("Performance por Lab")
    rep_labs_ids = df_labs_status[df_labs_status['name_rep'] == rep_name]['_id'].unique()
    rep_gatherings = df_gatherings_active[df_gatherings_active['_laboratory'].isin(rep_labs_ids)]
    
    if not rep_gatherings.empty:
        lab_performance = rep_gatherings.groupby('fantasyName').size().reset_index(name='Volume')
        lab_performance = lab_performance.sort_values('Volume', ascending=False)
        
        column_mapping = {
            'fantasyName': 'Laboratório',
            'Volume': 'Volume de Coletas'
        }
        lab_performance = lab_performance.rename(columns=column_mapping)
        
        # Adicionar dados
        for r in dataframe_to_rows(lab_performance, index=False, header=True):
            ws_performance.append(r)
        
        # Aplicar estilos
        for row in ws_performance.iter_rows(min_row=1, max_row=len(lab_performance)+1):
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
    
    # Aplicar layout e organização em todas as abas
    for ws in [ws_summary, ws_new, ws_descred, ws_cred, ws_status, ws_drops, ws_performance]:
        _autofit_freeze_and_filter(ws)

    # Salvar arquivo
    filename = f"relatorio_completo_{rep_name.replace(' ', '_')}_{current_date.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    return filename


def rep_individual_dashboard(rep_name: str, df_labs_status: pd.DataFrame, df_gatherings_active: pd.DataFrame, activity_window: int):
    """
    Dashboard individual por representante.
    """
    current_date = get_current_datetime()
    
    st.header(f"🎯 Análise Individual: {rep_name}")
    
    # Métricas de credenciamento
    accred_metrics = compute_rep_accreditations(df_labs_status, rep_name, current_date)
    
    # Status de atividade
    status_metrics = compute_rep_lab_status(df_labs_status, df_gatherings_active, rep_name, activity_window)
    
    # Pré-computar gráficos uma única vez para reutilizar (evita recomputo)
    charts = create_rep_charts(rep_name, df_gatherings_active, df_labs_status, activity_window)

    # Métricas principais
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Labs Credenciados", accred_metrics['num_credenciados'])
    with col2:
        st.metric("Labs Ativos", status_metrics['ativos'])
    with col3:
        st.metric("Novos Credenciamentos (3m)", accred_metrics['num_new_accred'])
    with col4:
        st.metric("Labs Inativos", status_metrics['inativos'])
    
    # Taxa de ativação
    total_labs = accred_metrics['num_credenciados']
    if total_labs > 0:
        activation_rate = (status_metrics['ativos'] / total_labs) * 100
        st.metric("Taxa de Ativação", f"{activation_rate:.1f}%")
    
    # Alertas e insights
    st.subheader("🚨 Alertas e Insights")
    
    alerts = []
    if status_metrics['inativos'] > 0:
        alerts.append(f"⚠️ {status_metrics['inativos']} labs inativos precisam de atenção")
    
    if accred_metrics['num_descred'] > 0:
        alerts.append(f"❌ {accred_metrics['num_descred']} labs descredenciados")
    
    if accred_metrics['num_new_accred'] == 0:
        alerts.append("📈 Nenhum novo credenciamento nos últimos 3 meses")
    
    if activation_rate < 70:
        alerts.append("📊 Taxa de ativação baixa - revisar estratégia")
    
    if alerts:
        for alert in alerts:
            st.warning(alert)
    else:
        st.success("✅ Performance excelente! Todos os indicadores estão positivos.")
    
    # Botões para exportar relatórios
    st.subheader("📄 Exportar Relatórios")
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        if st.button("📊 PDF Executivo", type="primary"):
            # Gerar PDF executivo
            try:
                # Preparar datasets reais
                # Top labs por volume
                rep_labs_ids = df_labs_status[df_labs_status['name_rep'] == rep_name]['_id'].unique()
                rep_g = df_gatherings_active[df_gatherings_active['_laboratory'].isin(rep_labs_ids)]
                top_labs = pd.DataFrame()
                if not rep_g.empty:
                    top_labs = rep_g.groupby('fantasyName').size().reset_index(name='Volume').sort_values('Volume', ascending=False)
                    top_labs = top_labs.rename(columns={'fantasyName': 'Laboratorio'})

                # Inativos criticos
                status_df = status_metrics.get('status_df', pd.DataFrame())
                inativos_pdf = pd.DataFrame()
                if not status_df.empty:
                    inativos_pdf = status_df[status_df['ativo_coleta'] == False].copy()
                    inativos_pdf = inativos_pdf.rename(columns={'fantasyName': 'Laboratorio', 'days_since_last_display': 'Dias sem Coletar', 'ultima_coleta_str': 'Ultima Coleta'})
                    if 'Dias sem Coletar' not in inativos_pdf.columns and 'days_since_last' in inativos_pdf.columns:
                        inativos_pdf['Dias sem Coletar'] = inativos_pdf['days_since_last']

                # Maiores quedas
                drops_pdf = detect_lab_drops(df_gatherings_active, rep_name)
                if not drops_pdf.empty:
                    drops_pdf = drops_pdf.sort_values('variation').copy()

                pdf_filename = generate_executive_pdf(rep_name, accred_metrics, status_metrics, charts, current_date, top_labs_df=top_labs, inactives_df=inativos_pdf, drops_df=drops_pdf)
                
                # Ler arquivo PDF para download
                with open(pdf_filename, "rb") as pdf_file:
                    pdf_bytes = pdf_file.read()
                
                # Botão de download
                st.download_button(
                    label="📥 Baixar PDF Executivo",
                    data=pdf_bytes,
                    file_name=pdf_filename,
                    mime="application/pdf"
                )
                
                # Limpar arquivo temporário
                import os
                os.remove(pdf_filename)
                
            except Exception as e:
                st.error(f"Erro ao gerar PDF: {str(e)}")
    
    with col2:
        if st.button("📋 Excel Completo", type="secondary"):
            try:
                # Gerar Excel completo
                excel_filename = generate_complete_excel(rep_name, accred_metrics, status_metrics, df_gatherings_active, df_labs_status, activity_window, current_date)
                
                # Ler arquivo Excel para download
                with open(excel_filename, "rb") as excel_file:
                    excel_bytes = excel_file.read()
                
                # Botão de download
                st.download_button(
                    label="📥 Baixar Excel Completo",
                    data=excel_bytes,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Limpar arquivo temporário
                import os
                os.remove(excel_filename)
                
            except Exception as e:
                st.error(f"Erro ao gerar Excel: {str(e)}")
    
    with col3:
        st.info("💡 PDF Executivo: Resumo de ações. Excel Completo: Todos os dados organizados em abas.")
    
    # Gráficos de análise
    st.subheader("📊 Análise Gráfica")
    
    # Explicação dos status
    with st.expander("ℹ️ Explicação dos Status", expanded=False):
        st.markdown("""
        **📋 Status de Atividade de Coleta:**
        - **🟢 Ativos em Coleta**: Laboratórios que realizaram coletas nos últimos **{} dias**
        - **🔴 Inativos em Coleta**: Laboratórios que **não coletaram** nos últimos **{} dias**
        
        **💡 Importante:** Este status é baseado na **atividade de coleta**, não no credenciamento. 
        Um laboratório pode estar credenciado mas inativo em coletas.
        """.format(activity_window, activity_window))
    
    if charts:
        # Layout em colunas para os gráficos
        col1, col2 = st.columns(2)
        
        with col1:
            if 'monthly_volume' in charts:
                st.plotly_chart(charts['monthly_volume'], use_container_width=True)
            
            if 'status_pie' in charts:
                st.plotly_chart(charts['status_pie'], use_container_width=True)
        
        with col2:
            if 'weekly_trend' in charts:
                st.plotly_chart(charts['weekly_trend'], use_container_width=True)
            
            if 'lab_performance' in charts:
                st.plotly_chart(charts['lab_performance'], use_container_width=True)
    else:
        st.info("📈 Nenhum dado disponível para gerar gráficos.")
    
    # Quedas em labs
    st.subheader("⚠️ Laboratórios com Quedas Bruscas")
    # Seleção de mês (padrão: mês atual)
    available_months = sorted(df_gatherings_active['createdAt'].dt.to_period('M').astype(str).unique()) if not df_gatherings_active.empty else []
    default_index = len(available_months) - 1 if available_months else 0
    selected_month = st.selectbox("Mês", options=available_months, index=default_index) if available_months else None
    original_drops = detect_lab_drops(df_gatherings_active, rep_name, target_month=selected_month)
    if not original_drops.empty and 'variation' in original_drops.columns:
        # Garantir mes como string para evitar problemas em exportacoes
        if 'month' in original_drops.columns and not pd.api.types.is_string_dtype(original_drops['month']):
            original_drops['month'] = original_drops['month'].astype(str)
        # Fazer uma cópia para formatação
        drops_df = original_drops.copy()
        drops_df['variation'] = drops_df['variation'].apply(lambda x: f"{x:.1f}%")
        drops_df['Volume'] = drops_df['Volume'].apply(lambda x: f"{int(x):,}".replace(",", "."))
        
        # Calcular mês anterior para referência
        drops_df['previous_month'] = drops_df['month'].apply(lambda x: str(pd.Period(x) - 1))
        drops_df['previous_volume'] = drops_df['previous_volume'].apply(lambda x: f"{int(x):,}".replace(",", "."))
        
        # Criar coluna com mês de referência
        drops_df['previous_volume_with_month'] = drops_df['previous_volume'] + ' (' + drops_df['previous_month'] + ')'
        
        display_columns = ['lab_info', 'month', 'Volume', 'previous_volume_with_month', 'variation']
        column_mapping = {
            'lab_info': 'Laboratório',
            'month': 'Mês',
            'Volume': 'Coletas Atual',
            'previous_volume_with_month': 'Coletas Anterior (Mês)',
            'variation': 'Queda (%)'
        }
        
        drops_df = drops_df[display_columns]
        drops_df = drops_df.rename(columns=column_mapping)
        
        # Destacar quedas mais críticas
        st.warning(f"🚨 {len(drops_df)} laboratórios com quedas bruscas detectadas!")
        st.dataframe(drops_df, use_container_width=True)
        
        # Resumo executivo
        st.subheader("📊 Resumo Executivo")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total de Labs com Quedas", len(drops_df))
        with col2:
            try:
                avg_drop = original_drops['variation'].mean()
                st.metric("Queda Média", f"{avg_drop:.1f}%")
            except (KeyError, ValueError):
                st.metric("Queda Média", "N/A")
        with col3:
            try:
                max_drop = original_drops['variation'].min()  # min() porque são valores negativos
                st.metric("Maior Queda", f"{max_drop:.1f}%")
            except (KeyError, ValueError):
                st.metric("Maior Queda", "N/A")
    else:
        st.success("✅ Nenhum laboratório com queda brusca detectada.")
    
    # Listas detalhadas - Dados limpos e organizados
    with st.expander("🆕 Novos Credenciamentos"):
        if not accred_metrics['new_accred_df'].empty:
            # Limpar e organizar dados de novos credenciamentos
            new_labs_clean = accred_metrics['new_accred_df'].copy()
            
            # Selecionar apenas colunas úteis para o gestor
            useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'data_credenciamento', 'dias_credenciado']
            available_columns = [col for col in useful_columns if col in new_labs_clean.columns]
            
            if available_columns:
                new_labs_clean = new_labs_clean[available_columns]
                
                # Renomear colunas para português
                column_mapping = {
                    'fantasyName': 'Nome do Laboratório',
                    'cnpj': 'CNPJ',
                    'name_rep': 'Representante',
                    'Categoria': 'Tipo',
                    'data_credenciamento': 'Data de Credenciamento',
                    'dias_credenciado': 'Dias Credenciado'
                }
                
                new_labs_clean = new_labs_clean.rename(columns=column_mapping)
                
                # Formatar CNPJ
                if 'CNPJ' in new_labs_clean.columns:
                    new_labs_clean['CNPJ'] = new_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                
                # Ordenar por data de credenciamento (mais recentes primeiro)
                if 'Data de Credenciamento' in new_labs_clean.columns:
                    new_labs_clean = new_labs_clean.sort_values('Data de Credenciamento', ascending=False)
                
                st.dataframe(new_labs_clean, use_container_width=True)
            else:
                st.info("Dados de novos credenciamentos não disponíveis.")
        else:
            st.info("Nenhum novo credenciamento encontrado.")
    
    with st.expander("❌ Descredenciados"):
        if not accred_metrics['descred_df'].empty:
            # Limpar e organizar dados de descredenciados
            descred_labs_clean = accred_metrics['descred_df'].copy()
            
            # Selecionar apenas colunas úteis
            useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'exclusionDate']
            available_columns = [col for col in useful_columns if col in descred_labs_clean.columns]
            
            if available_columns:
                descred_labs_clean = descred_labs_clean[available_columns]
                
                # Renomear colunas
                column_mapping = {
                    'fantasyName': 'Nome do Laboratório',
                    'cnpj': 'CNPJ',
                    'name_rep': 'Representante',
                    'Categoria': 'Tipo',
                    'exclusionDate': 'Data de Exclusão'
                }
                
                descred_labs_clean = descred_labs_clean.rename(columns=column_mapping)
                
                # Formatar CNPJ
                if 'CNPJ' in descred_labs_clean.columns:
                    descred_labs_clean['CNPJ'] = descred_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                
                st.dataframe(descred_labs_clean, use_container_width=True)
            else:
                st.info("Dados de descredenciados não disponíveis.")
        else:
            st.info("Nenhum laboratório descredenciado encontrado.")
    
    with st.expander("📋 Labs Credenciados"):
        if not accred_metrics['cred_df'].empty:
            # Limpar e organizar dados de credenciados
            cred_labs_clean = accred_metrics['cred_df'].copy()
            
            # Selecionar apenas colunas úteis
            useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'createdAt']
            available_columns = [col for col in useful_columns if col in cred_labs_clean.columns]
            
            if available_columns:
                cred_labs_clean = cred_labs_clean[available_columns]
                
                # Renomear colunas
                column_mapping = {
                    'fantasyName': 'Nome do Laboratório',
                    'cnpj': 'CNPJ',
                    'name_rep': 'Representante',
                    'Categoria': 'Tipo',
                    'createdAt': 'Data de Criação'
                }
                
                cred_labs_clean = cred_labs_clean.rename(columns=column_mapping)
                
                # Formatar CNPJ
                if 'CNPJ' in cred_labs_clean.columns:
                    cred_labs_clean['CNPJ'] = cred_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                
                # Ordenar por nome do laboratório
                cred_labs_clean = cred_labs_clean.sort_values('Nome do Laboratório')
                
                st.dataframe(cred_labs_clean, use_container_width=True)
            else:
                st.info("Dados de laboratórios credenciados não disponíveis.")
        else:
            st.info("Nenhum laboratório credenciado encontrado.")
    
    with st.expander("📊 Status de Coletas"):
        if not status_metrics['status_df'].empty:
            # Limpar e organizar dados de status
            status_labs_clean = status_metrics['status_df'].copy()
            
            # Selecionar apenas colunas úteis
            useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'ativo_coleta', 'ultima_coleta_str', 'days_since_last_display']
            available_columns = [col for col in useful_columns if col in status_labs_clean.columns]
            
            if available_columns:
                status_labs_clean = status_labs_clean[available_columns]
                
                # Renomear colunas
                column_mapping = {
                    'fantasyName': 'Nome do Laboratório',
                    'cnpj': 'CNPJ',
                    'name_rep': 'Representante',
                    'Categoria': 'Tipo',
                    'ativo_coleta': 'Ativo em Coletas',
                    'ultima_coleta_str': 'Última Coleta',
                    'days_since_last_display': 'Dias sem Coletar'
                }
                
                status_labs_clean = status_labs_clean.rename(columns=column_mapping)
                
                # Formatar CNPJ
                if 'CNPJ' in status_labs_clean.columns:
                    status_labs_clean['CNPJ'] = status_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                
                # Ordenar por status (ativos primeiro)
                if 'Ativo em Coletas' in status_labs_clean.columns:
                    status_labs_clean = status_labs_clean.sort_values('Ativo em Coletas', ascending=False)
                
                st.dataframe(status_labs_clean, use_container_width=True)
            else:
                st.info("Dados de status de coletas não disponíveis.")
        else:
            st.info("Nenhum dado de status de coletas encontrado.")


@st.cache_data
def load_data():
    df_reps, df_labs, df_gath = load_csvs()
    df_reps, df_labs = enrich_labs_with_reps(df_reps, df_labs)
    df_gath = merge_gatherings_with_labs(df_gath, df_labs)
    return df_reps, df_labs, df_gath


def create_search_options(df_labs: pd.DataFrame) -> list:
    options = ["Todos os laboratórios"]
    other_options = []
    seen = set()
    for _, row in df_labs.iterrows():
        fantasy_name = str(row.get('fantasyName', '') or '').strip()
        cnpj = str(row.get('cnpj', '') or '').strip()
        label = None
        if fantasy_name and fantasy_name != 'nan':
            label = f"{fantasy_name} (CNPJ: {cnpj})" if cnpj else fantasy_name
        elif cnpj and cnpj != 'nan' and cnpj != '00000000000000':
            label = f"CNPJ: {cnpj}"
        if label and label not in seen:
            seen.add(label)
            other_options.append(label)
    options.extend(sorted(other_options))
    return options


def main():
    st.set_page_config(page_title="Dashboard de Gestão Comercial - ToxRepresentatives", layout="wide")
    
    # Mudança: Adicionado CSS para fontes maiores e botões acessíveis.
    # Justificativa: Melhora usabilidade para usuários idosos; maintainable como inline style.
    st.markdown("""
        <style>
            .stApp {
                font-size: 18px;
            }
            button {
                height: 50px;
                font-size: 18px;
            }
            .stSelectbox, .stMultiselect, .stSlider {
                font-size: 18px;
            }
            .metric-label {
                font-size: 18px !important;
            }
            .metric-value {
                font-size: 24px !important;
            }
        </style>
    """, unsafe_allow_html=True)
    
    st.title("🏢 Dashboard de Gestão Comercial - ToxRepresentatives")

    current_date = get_current_datetime()

    # Dados
    df_reps, df_labs, df_gatherings = load_data()
    
    df_gatherings_merged = df_gatherings.copy()

    # Sidebar: filtros globais simplificados
    st.sidebar.header("🔍 Filtros Simples")
    # Verificar se createdAt é datetime e extrair anos
    if "createdAt" in df_gatherings.columns and pd.api.types.is_datetime64_any_dtype(df_gatherings["createdAt"]):
        year_options = [int(y) for y in sorted(df_gatherings["createdAt"].dt.year.dropna().unique())]
    else:
        year_options = [DEFAULT_YEAR]
    year = st.sidebar.selectbox("Ano", options=year_options, index=year_options.index(DEFAULT_YEAR) if DEFAULT_YEAR in year_options else 0)
    
    activity_window = st.sidebar.slider("Dias para Atividade", 7, 60, DEFAULT_ACTIVITY_WINDOW_DAYS, help="Define quando um lab é considerado ativo")
    tipo_rep = st.sidebar.selectbox("Tipo de Representante", ["Todos", "Interno", "Externo"])
    selected_reps = st.sidebar.multiselect("Representantes", sorted(df_labs["name_rep"].dropna().unique()))
    
    search_options = create_search_options(df_labs)
    selected_search = st.sidebar.selectbox("Buscar Lab ou CNPJ", options=search_options)
    
    if st.sidebar.button("Limpar Todos os Filtros", key="clear_filters"):
        st.session_state.clear()
        st.rerun()

    # Filtragem
    df_gatherings = df_gatherings[df_gatherings["createdAt"].dt.year == year]
    df_gatherings_active = filter_active_gatherings(df_gatherings, exclude_test=False, exclude_disabled=True)

    if tipo_rep != "Todos":
        df_labs = df_labs[df_labs["Categoria"] == tipo_rep]
        df_gatherings_active = df_gatherings_active[df_gatherings_active["Categoria"] == tipo_rep]

    if selected_reps:
        df_labs = df_labs[df_labs["name_rep"].isin(selected_reps)]
        lab_ids = df_labs["_id"].unique()
        df_gatherings_active = df_gatherings_active[df_gatherings_active["_laboratory"].isin(lab_ids)]

    if selected_search != "Todos os laboratórios":
        if selected_search.startswith("CNPJ:"):
            search_term = selected_search.split(" - ")[0].replace("CNPJ: ", "")
            mask = df_labs["cnpj"].astype(str).str.contains(search_term, case=False, na=False)
        else:
            search_term = selected_search.split(" (CNPJ:")[0]
            mask = df_labs["fantasyName"].astype(str).str.contains(search_term, case=False, na=False)
        df_labs = df_labs[mask]
        lab_ids = df_labs["_id"].unique()
        df_gatherings_active = df_gatherings_active[df_gatherings_active["_laboratory"].isin(lab_ids)]

    # Computações base
    df_labs_status = compute_credenciamento(df_labs, current_date)
    df_labs_cred = df_labs_status[df_labs_status["is_credenciado"]]
    credenciados = int(df_labs_cred.shape[0])
    descredenciados = int(df_labs_status.shape[0] - credenciados)

    df_labs_cred, ativos_coleta, inativos_coleta, last_collection = compute_coleta_status(
        df_labs_cred, df_gatherings_active, current_date, activity_window
    )

    weekly, monthly = aggregate_volumes(df_gatherings_active)
    kpis = compute_kpis(monthly)

    # Tabs reorganizadas: priorizar o que o gestor precisa
    # Mudança: Tabs reordenadas para facilitar acesso (visão geral e análises críticas primeiro).
    # Justificativa: UX otimizada para usuário sênior; maintainable com tabs claras.
    tab_dashboard, tab_rep_individual, tab_variations, tab_gestao, tab_alertas, tab_ranking, tab_geografia, tab_labs = st.tabs(
        ["📊 Visão Geral", "👤 Análise por Representante", "📉 Quedas Mensais", "🎯 Gestão Comercial", "⚠️ Alertas", "🏆 Ranking", "🗺️ Geografia", "🏥 Labs"]
    )

    with tab_dashboard:
        st.header("📊 Visão Geral Simples")
        kpi_cards(kpis, credenciados, descredenciados, ativos_coleta, inativos_coleta, activity_window)
        line_chart_monthly(monthly)
        line_chart_weekly(weekly)

    with tab_variations:
        st.header("📉 Análise de Quedas Mensais")
        st.info("Aqui você vê facilmente se houve queda entre meses e o quanto caiu.")
        # Se Tipo de Representante = Todos, condensar dados para evitar duplicações
        monthly_to_show = monthly
        if tipo_rep == "Todos" and not monthly.empty:
            monthly_to_show = monthly.groupby('month', as_index=False)['Volume'].sum()
        line_chart_with_variations(monthly_to_show)

    with tab_rep_individual:
        st.header("👤 Análise Individual de Representante")
        st.info("Selecione um representante para ver detalhes simples sobre seus labs, credenciamentos e quedas.")
        
        rep_options = sorted(df_labs["name_rep"].dropna().unique())
        selected_rep = st.selectbox("Escolha o Representante", options=rep_options)
        
        if selected_rep:
            rep_individual_dashboard(selected_rep, df_labs_status, df_gatherings_active, activity_window)

    with tab_gestao:
        rep_metrics = compute_representative_metrics(df_gatherings_active, df_labs_status, current_date, activity_window)
        category_summary = compute_category_summary(df_gatherings_active, df_labs_status)
        performance_dashboard(rep_metrics, category_summary)
        
        st.subheader("📊 Performance por Representante")
        representative_table(
            rep_metrics[['name_rep', 'Categoria', 'labs_credenciados', 'labs_ativos', 'labs_inativos', 'total_coletas', 'taxa_ativacao', 'produtividade']],
            "Performance"
        )
        
        st.subheader("🆕 Novos Credenciamentos")
        new_accred = compute_new_accreditations(df_labs_status, current_date, 3)
        
        if not new_accred.empty:
            # Limpar e organizar dados de novos credenciamentos
            new_accred_clean = new_accred.copy()
            
            # Selecionar apenas colunas úteis
            useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'data_credenciamento', 'dias_credenciado']
            available_columns = [col for col in useful_columns if col in new_accred_clean.columns]
            
            if available_columns:
                new_accred_clean = new_accred_clean[available_columns]
                
                # Renomear colunas
                column_mapping = {
                    'fantasyName': 'Nome do Laboratório',
                    'cnpj': 'CNPJ',
                    'name_rep': 'Representante',
                    'Categoria': 'Tipo',
                    'data_credenciamento': 'Data de Credenciamento',
                    'dias_credenciado': 'Dias Credenciado'
                }
                
                new_accred_clean = new_accred_clean.rename(columns=column_mapping)
                
                # Formatar CNPJ
                if 'CNPJ' in new_accred_clean.columns:
                    new_accred_clean['CNPJ'] = new_accred_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                
                # Ordenar por data de credenciamento (mais recentes primeiro)
                if 'Data de Credenciamento' in new_accred_clean.columns:
                    new_accred_clean = new_accred_clean.sort_values('Data de Credenciamento', ascending=False)
                
                st.dataframe(new_accred_clean, use_container_width=True)
        else:
            st.info("Dados de novos credenciamentos não disponíveis.")


    with tab_alertas:
        st.header("⚠️ Alertas - Laboratórios Inativos")
        threshold_days = st.slider("Dias para Alerta de Inatividade", 15, 90, 30, help="Define quantos dias sem coleta para considerar um lab inativo")
        
        inactive_labs = compute_inactive_labs_alert(df_labs_status, df_gatherings_merged, current_date, threshold_days)
        
        if not inactive_labs.empty:
            # Resumo executivo
            st.warning(f"🚨 {len(inactive_labs)} laboratórios inativos há mais de {threshold_days} dias!")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total de Labs Inativos", len(inactive_labs))
            with col2:
                avg_days = inactive_labs['dias_sem_coletar'].mean() if 'dias_sem_coletar' in inactive_labs.columns else 0
                st.metric("Média de Dias Inativo", f"{avg_days:.0f} dias")
            with col3:
                max_days = inactive_labs['dias_sem_coletar'].max() if 'dias_sem_coletar' in inactive_labs.columns else 0
                st.metric("Mais Tempo Inativo", f"{max_days:.0f} dias")
            
            # Agrupar por representante para análise
            if 'name_rep' in inactive_labs.columns:
                st.subheader("📋 Análise por Representante")
                rep_summary = inactive_labs.groupby('name_rep').size().reset_index(name='labs_inativos')
                rep_summary = rep_summary.sort_values('labs_inativos', ascending=False)
                
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.dataframe(rep_summary.rename(columns={
                        'name_rep': 'Representante',
                        'labs_inativos': 'Labs Inativos'
                    }), use_container_width=True)
                
                with col2:
                    st.subheader("🎯 Ações Recomendadas")
                    st.markdown("""
                    - **Contatar representantes** com mais labs inativos
                    - **Verificar motivos** da inatividade
                    - **Planejar visitas** aos labs críticos
                    - **Revisar estratégia** de relacionamento
                    """)
            
            # Lista detalhada
            st.subheader("📋 Lista Detalhada de Labs Inativos")
            
            if not inactive_labs.empty:
                # Limpar e organizar dados de labs inativos
                inactive_labs_clean = inactive_labs.copy()
                
                # Selecionar apenas colunas úteis
                useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'ultima_coleta_str', 'dias_sem_coletar_display']
                available_columns = [col for col in useful_columns if col in inactive_labs_clean.columns]
                
                if available_columns:
                    inactive_labs_clean = inactive_labs_clean[available_columns]
                    
                    # Renomear colunas
                    column_mapping = {
                        'fantasyName': 'Nome do Laboratório',
                        'cnpj': 'CNPJ',
                        'name_rep': 'Representante',
                        'Categoria': 'Tipo',
                        'ultima_coleta_str': 'Última Coleta',
                        'dias_sem_coletar_display': 'Dias sem Coletar'
                    }
                    
                    inactive_labs_clean = inactive_labs_clean.rename(columns=column_mapping)
                    
                    # Formatar CNPJ
                    if 'CNPJ' in inactive_labs_clean.columns:
                        inactive_labs_clean['CNPJ'] = inactive_labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                    
                    # Ordenar por dias sem coletar (mais críticos primeiro)
                    if 'Dias sem Coletar' in inactive_labs_clean.columns:
                        inactive_labs_clean = inactive_labs_clean.sort_values('Dias sem Coletar', ascending=False)
                    
                    st.dataframe(inactive_labs_clean, use_container_width=True)
                else:
                    st.info("Dados de labs inativos não disponíveis.")
            else:
                st.info("Nenhum lab inativo encontrado.")
        else:
            st.success(f"✅ Todos os laboratórios estão ativos! (coletaram nos últimos {threshold_days} dias)")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Status", "✅ Ativo")
            with col2:
                st.metric("Labs Ativos", len(df_labs_status))
            with col3:
                st.metric("Taxa de Ativação", "100%")

    # Outras tabs mantidas como antes, mas com fontes maiores implicitamente pelo CSS

    with tab_ranking:
        ranking_reps, ranking_labs = build_rankings(df_gatherings_active, df_labs_cred, last_collection, current_date, activity_window)
        st.subheader("🏆 Ranking de Representantes")
        table(ranking_reps, "Ranking Reps")
        st.subheader("🏆 Ranking de Laboratórios")
        table(ranking_labs, "Ranking Labs")

    

    with tab_geografia:
        state_metrics = compute_geographic_metrics(df_gatherings_active, df_labs_status, current_date, activity_window)
        city_metrics = compute_city_metrics(df_gatherings_active, df_labs_status, current_date, activity_window)
        geographic_dashboard(state_metrics, city_metrics)

    with tab_labs:
        st.header("🏥 Gestão de Laboratórios")
        if not df_labs_cred.empty:
            labs_clean = df_labs_cred.copy()
            useful_columns = ['fantasyName', 'cnpj', 'name_rep', 'Categoria', 'ativo_coleta', 'ultima_coleta_str', 'days_since_last_display']
            available_columns = [col for col in useful_columns if col in labs_clean.columns]
            if available_columns:
                labs_clean = labs_clean[available_columns]
                column_mapping = {
                    'fantasyName': 'Nome do Laboratório',
                    'cnpj': 'CNPJ',
                    'name_rep': 'Representante',
                    'Categoria': 'Tipo',
                    'ativo_coleta': 'Ativo em Coletas',
                    'ultima_coleta_str': 'Última Coleta',
                    'days_since_last_display': 'Dias sem Coletar'
                }
                labs_clean = labs_clean.rename(columns=column_mapping)
                if 'CNPJ' in labs_clean.columns:
                    labs_clean['CNPJ'] = labs_clean['CNPJ'].apply(lambda x: f"{x[:2]}.{x[2:5]}.{x[5:8]}/{x[8:12]}-{x[12:]}" if pd.notna(x) and len(str(x)) == 14 else str(x))
                labs_clean = labs_clean.sort_values('Nome do Laboratório')
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total de Labs", len(labs_clean))
                with col2:
                    ativos = len(labs_clean[labs_clean['Ativo em Coletas'] == True]) if 'Ativo em Coletas' in labs_clean.columns else 0
                    st.metric("Labs Ativos", ativos)
                with col3:
                    inativos = len(labs_clean[labs_clean['Ativo em Coletas'] == False]) if 'Ativo em Coletas' in labs_clean.columns else 0
                    st.metric("Labs Inativos", inativos)
                st.dataframe(labs_clean, use_container_width=True)
            else:
                st.info("Dados de laboratórios não disponíveis.")
        else:
            st.info("Nenhum laboratório encontrado.")


if __name__ == "__main__":
    main()

